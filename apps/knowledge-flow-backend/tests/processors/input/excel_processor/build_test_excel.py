# Copyright Thales 2025
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
build_test_excel.py
===================
Construit un classeur Excel de test volontairement « sale » pour exercer
toutes les branches de l'extracteur (`excel_extractor.ExcelExtractor`) :
colonnes/lignes masquées, tableaux empilés, colonnes-étiquettes, orientations
transposée/croisée, typage, en-têtes dupliqués, formules/erreurs, résidus,
découpe et compaction.

Usage :
    python build_test_excel.py [chemin_de_sortie.xlsx]
"""

import sys

from openpyxl import Workbook


# --------------------------------------------------------------------------- #
# Démonstration : on fabrique un classeur volontairement « sale »
# --------------------------------------------------------------------------- #
def build_demo_workbook(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"

    # Tableau 1 : colonnes A-G, colonne C et F masquées (colonnes internes cachées)
    # Résultat attendu avec include_hidden_cells=False : A, B, D, E, G seulement
    ws["A1"] = "Produit"
    ws["B1"] = "Ventes_2023"
    ws["C1"] = "CACHE_formule_interne"  # colonne C masquée
    ws["D1"] = "Ventes_2024"
    ws["E1"] = "Croissance"
    ws["F1"] = "CACHE_id_interne"  # colonne F masquée
    ws["G1"] = "Statut"
    for i, row in enumerate(
        [
            ["Stylo", 100, 9999, 130, 30, "ID001", "ok"],
            ["Cahier", 90, 8888, 95, 5, "ID002", "ok"],
            ["Stylet", 60, 7777, 80, 20, "ID003", "nok"],
        ],
        start=2,
    ):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["F"].hidden = True

    # Séparateur
    # Tableau 2 : ligne 6 masquée (sous-total intermédiaire caché)
    ws["A6"], ws["B6"], ws["C6"], ws["D6"] = "Région", "Ville", "Ventes", "Code"
    ws["A7"], ws["B7"], ws["C7"], ws["D7"] = "Nord", "Lille", 1250, "0042"
    ws["A8"], ws["B8"], ws["C8"], ws["D8"] = "Nord", "Lens", 1180, "0043"
    ws["A9"], ws["B9"], ws["C9"], ws["D9"] = "CACHE", "SOUS-TOTAL", 2430, ""  # ligne masquée
    ws["A10"], ws["B10"], ws["C10"], ws["D10"] = "Sud", "Nice", 980, "0100"
    ws.row_dimensions[9].hidden = True

    # Feuille « Empilés » : deux tableaux collés haut→bas, séparés par une
    # ligne-titre fusionnée pleine largeur (PAS de ligne vide entre eux).
    # → une seule composante connexe qui doit être re-découpée en 2 tableaux,
    #   le second portant en plus une ligne de contexte.
    ws3 = wb.create_sheet("Empilés")
    ws3["A1"] = "Tableau A — ventes trimestrielles"  # titre du 1er tableau
    ws3.merge_cells("A1:D1")
    ws3["A2"], ws3["B2"], ws3["C2"], ws3["D2"] = "Produit", "Q1", "Q2", "Q3"
    ws3["A3"], ws3["B3"], ws3["C3"], ws3["D3"] = "Stylo", 100, 110, 120
    ws3["A4"], ws3["B4"], ws3["C4"], ws3["D4"] = "Cahier", 90, 95, 100
    ws3["A5"] = "Tableau B — stocks"  # titre du 2e tableau
    ws3.merge_cells("A5:D5")
    ws3["A6"] = "au 31/12/2024, tous entrepôts"  # contexte du 2e tableau
    ws3.merge_cells("A6:D6")
    ws3["A7"], ws3["B7"], ws3["C7"], ws3["D7"] = "Produit", "Entrepôt", "Quantité", "Seuil"
    ws3["A8"], ws3["B8"], ws3["C8"], ws3["D8"] = "Stylo", "Lille", 500, 50
    ws3["A9"], ws3["B9"], ws3["C9"], ws3["D9"] = "Cahier", "Lens", 300, 30

    # Feuille « Séparateur 80% » : deux tableaux empilés séparés par une ligne-titre
    # fusionnée NON pleine largeur (A:D sur 5 colonnes = 80%). L'ancien algo exigeait
    # une fusion pleine largeur (A:E) et n'aurait PAS découpé ; avec
    # SPLIT_SEPARATOR_WIDTH_RATIO=0.8 la fusion A:D suffit → 2 tableaux attendus.
    # La colonne E (« Total »), hors de la fusion-séparateur, reste une vraie colonne.
    ws4 = wb.create_sheet("Séparateur 80%")
    ws4["A1"] = "Ventes Q1 — zone principale"  # titre 1 (fusion A:D = 80%)
    ws4.merge_cells("A1:D1")
    ws4["A2"], ws4["B2"], ws4["C2"], ws4["D2"], ws4["E2"] = "Produit", "Jan", "Fév", "Mar", "Total"
    ws4["A3"], ws4["B3"], ws4["C3"], ws4["D3"], ws4["E3"] = "Stylo", 10, 12, 11, 33
    ws4["A4"], ws4["B4"], ws4["C4"], ws4["D4"], ws4["E4"] = "Cahier", 8, 9, 7, 24
    ws4["A5"] = "Ventes Q2 — zone principale"  # titre 2 (fusion A:D = 80%)
    ws4.merge_cells("A5:D5")
    ws4["A6"], ws4["B6"], ws4["C6"], ws4["D6"], ws4["E6"] = "Produit", "Avr", "Mai", "Jun", "Total"
    ws4["A7"], ws4["B7"], ws4["C7"], ws4["D7"], ws4["E7"] = "Stylo", 13, 14, 12, 39
    ws4["A8"], ws4["B8"], ws4["C8"], ws4["D8"], ws4["E8"] = "Cahier", 9, 10, 8, 27

    # Feuille « Colonnes-étiquettes » : deux tableaux empilés (séparés par leur
    # ligne-titre fusionnée pleine largeur, comme « Empilés »), chacun précédé de
    # colonnes-étiquettes fusionnées verticalement sur toute la hauteur du corps.
    # A5 doit les retirer (de gauche à droite) et verser leur valeur au contexte :
    #   • Tableau 1 → 1 colonne-étiquette (« Papeterie »).
    #   • Tableau 2 → 2 colonnes-étiquettes consécutives (« Nord », « Lille »),
    #                 puis arrêt à la 1re colonne non fusionnée (« Produit »).
    ws5 = wb.create_sheet("Colonnes-étiquettes")
    # — Tableau 1 : 1 colonne-étiquette —
    ws5["A1"] = "Catalogue Papeterie 2024"  # titre (pleine largeur A:E)
    ws5.merge_cells("A1:E1")
    ws5["A2"] = "Papeterie"  # étiquette verticale (corps rows 2-6)
    ws5.merge_cells("A2:A6")
    ws5["B2"], ws5["C2"], ws5["D2"], ws5["E2"] = "Produit", "Prix", "Stock", "Statut"
    ws5["B3"], ws5["C3"], ws5["D3"], ws5["E3"] = "Stylo", 1.5, 500, "ok"
    ws5["B4"], ws5["C4"], ws5["D4"], ws5["E4"] = "Cahier", 2.0, 300, "ok"
    ws5["B5"], ws5["C5"], ws5["D5"], ws5["E5"] = "Gomme", 0.8, 150, "ok"
    ws5["B6"], ws5["C6"], ws5["D6"], ws5["E6"] = "Règle", 1.2, 80, "nok"
    # — Tableau 2 (collé dessous) : 2 colonnes-étiquettes —
    ws5["A7"] = "Stock par dépôt"  # titre (pleine largeur A:E)
    ws5.merge_cells("A7:E7")
    ws5["A8"] = "Nord"  # étiquette 1 (corps rows 8-12)
    ws5.merge_cells("A8:A12")
    ws5["B8"] = "Lille"  # étiquette 2 (corps rows 8-12)
    ws5.merge_cells("B8:B12")
    ws5["C8"], ws5["D8"], ws5["E8"] = "Produit", "Quantité", "Seuil"
    ws5["C9"], ws5["D9"], ws5["E9"] = "Stylo", 500, 50
    ws5["C10"], ws5["D10"], ws5["E10"] = "Cahier", 300, 30
    ws5["C11"], ws5["D11"], ws5["E11"] = "Gomme", 120, 20
    ws5["C12"], ws5["D12"], ws5["E12"] = "Règle", 90, 10

    # Feuille « Transposée » : en-tête NUMÉRIQUE (années) + 1re colonne TEXTE
    # (indicateurs), corps entièrement numérique et tableau plus large que haut
    # (lignes ≤ colonnes+1). B1 doit détecter l'orientation « transposée » et
    # remettre le tableau d'aplomb (df.T) : années en lignes, indicateurs en
    # colonnes après transposition.
    ws6 = wb.create_sheet("Transposée")
    ws6["A1"], ws6["B1"], ws6["C1"], ws6["D1"], ws6["E1"] = "Indicateur", 2021, 2022, 2023, 2024
    ws6["A2"], ws6["B2"], ws6["C2"], ws6["D2"], ws6["E2"] = "Chiffre d'affaires", 100, 120, 150, 170
    ws6["A3"], ws6["B3"], ws6["C3"], ws6["D3"], ws6["E3"] = "Charges", 80, 90, 100, 110
    ws6["A4"], ws6["B4"], ws6["C4"], ws6["D4"], ws6["E4"] = "Résultat", 20, 30, 50, 60

    # Feuille « Croisée » : en-tête numérique + 1re colonne texte (comme la
    # transposée) MAIS tableau trop HAUT (lignes > colonnes+1) → la condition de
    # transposition échoue, B1 doit classer en « croisée ».
    ws7 = wb.create_sheet("Croisée")
    ws7["A1"], ws7["B1"], ws7["C1"] = "Mois", 2022, 2023
    for i, (mois, a, b) in enumerate(
        [
            ("Janvier", 100, 110),
            ("Février", 90, 95),
            ("Mars", 120, 130),
            ("Avril", 80, 85),
            ("Mai", 140, 150),
            ("Juin", 100, 105),
        ],
        start=2,
    ):
        ws7.cell(row=i, column=1, value=mois)
        ws7.cell(row=i, column=2, value=a)
        ws7.cell(row=i, column=3, value=b)

    # Feuille « Typage » : exerce les branches de reconnaissance B4
    # (à condition de lancer le pipeline avec coerce_types=True) :
    #   • Prix        → numérique : € + virgule décimale conservés (chaîne)
    #   • Remise      → numérique : % conservé (chaîne, pas de ÷100)
    #   • Quantité    → numérique : espaces de milliers retirés (chaîne)
    #   • Disponible  → texte (oui/non ; plus de branche booléenne)
    #   • Date_MAJ    → date en texte (jj/mm/aaaa, dayfirst)
    #   • Code_Postal → numérique reconnu, zéros de tête préservés (chaîne)
    #   • Téléphone   → numérique reconnu, zéros de tête préservés (chaîne)
    ws8 = wb.create_sheet("Typage")
    ws8["A1"], ws8["B1"], ws8["C1"], ws8["D1"], ws8["E1"], ws8["F1"], ws8["G1"], ws8["H1"] = (
        "Produit",
        "Prix",
        "Remise",
        "Quantité",
        "Disponible",
        "Date_MAJ",
        "Code_Postal",
        "Téléphone",
    )
    for i, row in enumerate(
        [
            ["Stylo", "1,50 €", "15%", "1 200", "oui", "01/03/2024", "01000", "0612345678"],
            ["Cahier", "2,00 €", "5%", "12 000", "non", "15/06/2024", "75001", "0698765432"],
            ["Gomme", "0,80 €", "0%", "8 500", "oui", "31/12/2023", "06000", "0700000000"],
        ],
        start=2,
    ):
        for j, v in enumerate(row, start=1):
            ws8.cell(row=i, column=j, value=v)

    # Feuille « Entêtes » : noms de colonnes dupliqués (« Montant ») et un nom
    # VIDE. B2 doit dédupliquer (« Montant », « Montant.1 », « Montant.2 ») et
    # nommer la colonne vide (« col_2 »).
    ws9 = wb.create_sheet("Entêtes")
    ws9["A1"], ws9["B1"], ws9["C1"], ws9["D1"] = "Montant", "Montant", None, "Montant"
    ws9["A2"], ws9["B2"], ws9["C2"], ws9["D2"] = 10, 20, 30, 40
    ws9["A3"], ws9["B3"], ws9["C3"], ws9["D3"] = 11, 21, 31, 41
    ws9["A4"], ws9["B4"], ws9["C4"], ws9["D4"] = 12, 22, 32, 42

    # Feuille « Formules » : colonne Total en FORMULES (=B*C) → A1 doit voir
    # has_formulas=True ; et trois cellules d'ERREUR littérales (#DIV/0!, #REF!,
    # #NAME?) → A2 doit toutes les capter (regex ^#.*[!?]$).
    ws10 = wb.create_sheet("Formules")
    ws10["A1"], ws10["B1"], ws10["C1"], ws10["D1"] = "Produit", "Quantité", "Prix", "Total"
    ws10["A2"], ws10["B2"], ws10["C2"], ws10["D2"] = "Stylo", 10, 2, "=B2*C2"
    ws10["A3"], ws10["B3"], ws10["C3"], ws10["D3"] = "Cahier", 5, 3, "=B3*C3"
    ws10["A4"], ws10["B4"], ws10["C4"], ws10["D4"] = "Division", 0, 0, "#DIV/0!"
    ws10["A5"], ws10["B5"], ws10["C5"], ws10["D5"] = "Référence", 1, 1, "#REF!"
    ws10["A6"], ws10["B6"], ws10["C6"], ws10["D6"] = "Nom", 2, 2, "#NAME?"

    # Feuille « Résidus » : un vrai tableau, PLUS deux blocs à classer en résidu
    # par A3 — un bloc non tabulaire (sous-total 2 cellules sur 1 ligne) et une
    # cellule isolée. Séparés du tableau par des lignes vides.
    ws11 = wb.create_sheet("Résidus")
    ws11["A1"], ws11["B1"], ws11["C1"] = "Produit", "Ventes", "Région"
    ws11["A2"], ws11["B2"], ws11["C2"] = "Stylo", 100, "Nord"
    ws11["A3"], ws11["B3"], ws11["C3"] = "Cahier", 90, "Sud"
    ws11["A4"], ws11["B4"], ws11["C4"] = "Gomme", 50, "Est"
    ws11["A7"], ws11["B7"] = "Sous-total", 240  # bloc_non_tabulaire (h<2)
    ws11["E10"] = "note libre non auditée"  # cellule_isolée

    # Feuille « Découpe » : exerce les résidus de DÉCOUPE A4 (visibles comme
    # résidus avec keep_split_residuals=True, sinon journalisés). Bloc 1 : un
    # tableau suivi d'une ligne-titre fusionnée FINALE sans corps → titre
    # orphelin. Bloc 2 : ligne-titre fusionnée suivie d'un seul en-tête (corps
    # 1 ligne) → corps trop maigre.
    ws12 = wb.create_sheet("Découpe")
    ws12["A1"], ws12["B1"], ws12["C1"], ws12["D1"] = "Produit", "Q1", "Q2", "Q3"
    ws12["A2"], ws12["B2"], ws12["C2"], ws12["D2"] = "Stylo", 1, 2, 3
    ws12["A3"], ws12["B3"], ws12["C3"], ws12["D3"] = "Cahier", 4, 5, 6
    ws12["A4"] = "Total annuel — note de bas de tableau"  # titre fusionné FINAL → orphelin
    ws12.merge_cells("A4:D4")
    # bloc 2, séparé par la ligne 5 vide
    ws12["A6"] = "Section sans données"  # titre fusionné…
    ws12.merge_cells("A6:D6")
    ws12["A7"], ws12["B7"], ws12["C7"], ws12["D7"] = "Produit", "Q1", "Q2", "Q3"  # …+ corps 1 ligne → maigre

    # Feuille « Compaction » : un tableau dont une ligne n'a QUE sa première
    # colonne remplie. B3 la conserve en mode "all" mais la supprime en mode
    # "except_first_col" → permet de distinguer les deux stratégies.
    ws13 = wb.create_sheet("Compaction")
    ws13["A1"], ws13["B1"], ws13["C1"] = "Catégorie", "Valeur1", "Valeur2"
    ws13["A2"], ws13["B2"], ws13["C2"] = "Alpha", 10, 20
    ws13["A3"], ws13["B3"], ws13["C3"] = "Beta", 30, 40
    ws13["A4"] = "Libellé orphelin"  # B4/C4 vides → 1re col. seule
    ws13["A5"], ws13["B5"], ws13["C5"] = "Gamma", 50, 60

    # Feuille masquée
    ws2 = wb.create_sheet("Brouillon")
    ws2.sheet_state = "hidden"
    ws2.append(["clé", "valeur"])
    ws2.append(["alpha", 1])
    ws2.append(["beta", 2])

    wb.save(path)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "test_book.xlsx"
    build_demo_workbook(out)
    print(f"OK classeur de test → {out}")
