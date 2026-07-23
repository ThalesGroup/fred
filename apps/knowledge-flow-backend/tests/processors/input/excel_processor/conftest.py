# Copyright Thales 2025
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Fixtures partagées des tests du processeur Excel.

Quatre niveaux s'appuient sur ce fichier :
  • test_helpers.py     — fonctions pures (helpers).
  • test_steps.py       — étapes du pipeline prises isolément (A1…B5).
  • test_integration.py — extraction complète sur le classeur de démo.
  • test_export.py      — export (output.md + extraits + tables.json).

Le processeur ne porte plus d'état par-fichier : chaque fixture construit un
`ExcelProcessor` configuré (options en attributs d'instance) puis :
  • `make_extractor` renvoie l'`ExcelExtractor` câblé par le vrai
    `ExcelProcessor._build_extractor` (étapes A/B) ;
  • `run_export` appelle le point d'entrée public `convert_file_to_markdown`
    (output.md + extraits + tables.json) ;
  • `demo_run` exécute l'extraction complète sur le classeur de démo (recalc
    LibreOffice) — l'ensemble est skippé quand LibreOffice est absent.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook

from knowledge_flow_backend.core.processors.input.excel_processor import excel_extractor as ee
from knowledge_flow_backend.core.processors.input.excel_processor.excel_processor import ExcelProcessor
from tests.processors.input.excel_processor import build_test_excel as btx


# --------------------------------------------------------------------------- #
# Constructeur de petits classeurs ciblés (tests unitaires d'étapes)
# --------------------------------------------------------------------------- #
def build_workbook(path, sheets):
    """Écrit un classeur xlsx à partir d'une description compacte.

    `sheets` : liste de dicts décrivant chaque feuille :
        {
          "name": "Feuille1",
          "cells": {"A1": "x", "B1": 2, ...},   # coord Excel -> valeur
          "merges": ["A1:D1", ...],             # optionnel
          "hidden_cols": ["C", "F"],            # optionnel
          "hidden_rows": [9],                    # optionnel (1-based)
          "state": "visible" | "hidden",        # optionnel
          "number_formats": {"A1": "0;\\-0;;"}, # optionnel
          "show_zeros": True | False,            # optionnel (réglage feuille)
        }
    Retourne `path` (str) pour enchaîner.
    """
    wb = Workbook()
    # On retire la feuille par défaut une fois la première créée.
    default = wb.active
    first = True
    for spec in sheets:
        if first:
            ws = default
            ws.title = spec["name"]
            first = False
        else:
            ws = wb.create_sheet(spec["name"])
        for coord, value in spec.get("cells", {}).items():
            ws[coord] = value
        for fmt_coord, fmt in spec.get("number_formats", {}).items():
            ws[fmt_coord].number_format = fmt
        for rng in spec.get("merges", []):
            ws.merge_cells(rng)
        for col in spec.get("hidden_cols", []):
            ws.column_dimensions[col].hidden = True
        for row in spec.get("hidden_rows", []):
            ws.row_dimensions[row].hidden = True
        if "show_zeros" in spec:
            ws.sheet_view.showZeros = spec["show_zeros"]
        ws.sheet_state = spec.get("state", "visible")
    wb.save(str(path))
    return str(path)


def _configure_processor(**cfg) -> ExcelProcessor:
    """Fresh processor with per-test overrides applied as instance attributes.

    Options are class attributes on `ExcelProcessor`; setting them on the
    instance overrides the defaults for one run. `output_dir` is no longer a
    config value (it is a `convert_file_to_markdown` argument), so it is dropped.
    """
    cfg.pop("output_dir", None)
    proc = ExcelProcessor()
    for key, value in cfg.items():
        setattr(proc, key, value)
    return proc


@pytest.fixture
def make_extractor(tmp_path):
    """Fabrique un `ExcelExtractor` configuré à partir d'une description de feuilles.

    L'extracteur est câblé exactement comme le fait le processeur
    (`ExcelProcessor._build_extractor`), donc les tests d'étapes exercent le vrai
    chemin de chargement/configuration. Tout kwarg supplémentaire devient une
    option du processeur (ex. `sheets`, `include_hidden_cells`,
    `apply_format_masking`, `split_on_hidden_columns`).

    Les méthodes d'extraction (a1…b5) vivent sur l'objet renvoyé.
    """
    counter = {"n": 0}

    def _factory(sheets_spec, **cfg):
        counter["n"] += 1
        path = tmp_path / f"wb_{counter['n']}.xlsx"
        build_workbook(path, sheets_spec)
        return _configure_processor(**cfg)._build_extractor(str(path))

    return _factory


@pytest.fixture
def make_table():
    """Construit un `DetectedTable` à la main pour les tests des étapes B.

    `grid_rows` : liste de listes (devient une grille numpy dtype=object).
    """

    def _factory(grid_rows, *, merges=None, sheet="S", bbox=None, **kwargs):
        grid = np.array(grid_rows, dtype=object)
        if bbox is None:
            bbox = (0, 0, grid.shape[0] - 1, grid.shape[1] - 1)
        return ee.DetectedTable(
            id=kwargs.pop("id", f"{sheet}.t1"),
            sheet=sheet,
            bbox=bbox,
            grid=grid,
            local_merges=list(merges or []),
            **kwargs,
        )

    return _factory


@pytest.fixture
def run_export(tmp_path):
    """Exécute la conversion complète sur un petit classeur et renvoie le résultat.

    Produit `output.md`, les extraits par table et `tables.json` sous
    `<tmp>/output_<n>` via le point d'entrée public `convert_file_to_markdown`.
    L'objet renvoyé expose `.proc`, `.path`, `.output_dir` (Path) et `.result`.
    """
    counter = {"n": 0}

    def _factory(sheets_spec, **cfg):
        counter["n"] += 1
        path = tmp_path / f"wb_{counter['n']}.xlsx"
        build_workbook(path, sheets_spec)
        proc = _configure_processor(**cfg)
        out_dir = tmp_path / f"output_{counter['n']}"
        document_uid = f"uid-{counter['n']}"
        result = proc.convert_file_to_markdown(Path(path), out_dir, document_uid)

        class _Run:
            pass

        r = _Run()
        r.proc = proc
        r.path = Path(path)
        r.output_dir = out_dir
        r.document_uid = document_uid
        r.result = result
        return r

    return _factory


# --------------------------------------------------------------------------- #
# Classeur de démo (intégration) — construit et traité une seule fois
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="session")
def demo_path(tmp_path_factory):
    """Construit le classeur de démo « sale » une fois pour toute la session."""
    path = tmp_path_factory.mktemp("demo") / "demo.xlsx"
    btx.build_demo_workbook(str(path))
    return str(path)


@pytest.fixture(scope="session")
def demo_run(tmp_path_factory, demo_path):
    """Exécute le pipeline complet sur le classeur de démo (config canonique).

    Retourne un objet avec :
      • .summaries        — liste de SheetSummary
      • .by_name          — dict {nom de feuille: SheetSummary}
      • .output_dir       — dossier où l'export a été produit (str)

    La config (include_hidden_cells=False, coerce_types=True) correspond aux
    comportements attendus documentés dans build_test_excel.py. `recalc=True`
    ré-évalue la feuille « Formules » via LibreOffice headless : tout le
    périmètre d'intégration est skippé quand LibreOffice est absent.
    """
    if shutil.which("soffice") is None:
        pytest.skip("LibreOffice requis pour le run d'intégration Excel (recalc=True)")

    out_dir = tmp_path_factory.mktemp("demo_out")
    proc = ExcelProcessor()
    proc.include_hidden_cells = False
    proc.recalc = True
    proc.split_on_hidden_columns = True
    # Le run canonique documenté exporte des extraits CSV (le défaut plateforme
    # est parquet, dédié au circuit d'ingestion SQL).
    proc.extract_format = "csv"
    # Feuille « Entêtes » (build_test_excel.py) documente le nommage de la
    # colonne vide en « col_2 » plutôt que sa suppression -> keep_headerless_columns=True.
    # ExcelProcessor la met à False par défaut (option produit conservatrice) ;
    # le run canonique documenté l'active pour exercer ce comportement de B2.
    proc.keep_headerless_columns = True
    # Feuille « Découpe » documente le cas keep_split_residuals=False (résidus
    # de découpe journalisés, non conservés dans le résumé). ExcelProcessor met
    # keep_split_residuals à True par défaut ; le run canonique documenté le
    # désactive pour exercer ce cas.
    proc.keep_split_residuals = False
    # include_hidden_sheets, coerce_types, keep_headerless_tables,
    # keep_single_column_tables et show_column_names gardent leurs valeurs par
    # défaut d'ExcelProcessor : aucune feuille de démo n'exerce ces cas.

    extractor = proc._build_extractor(demo_path)
    summaries = extractor.extract()
    proc._export(extractor.all_tables, summaries, out_dir)

    class _Result:
        pass

    r = _Result()
    r.summaries = summaries
    r.by_name = {s.name: s for s in summaries}
    r.output_dir = str(out_dir)
    r.proc = proc
    return r
