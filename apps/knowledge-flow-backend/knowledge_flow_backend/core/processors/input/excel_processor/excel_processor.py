# Copyright Thales 2025
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
excel_processor.py
==================
Excel input processor: hands a workbook to the extractor, then writes the preview.

``ExcelProcessor`` is a :class:`BaseMarkdownProcessor`. It exposes the standard
``check_file_validity`` / ``extract_file_metadata`` / ``convert_file_to_markdown``
entry points and internally drives an :class:`~excel_extractor.ExcelExtractor`:
it checks the source file exists, optionally recalculates it with headless
LibreOffice (then deletes the temporary recalculated file), loads the
value/formula workbooks, runs phases A and B via the extractor, and finally
generates the output files (per-table extracts + Markdown summary).

The processing RULES themselves (phases A and B) and the shared data structures
/ helpers live in ``excel_extractor``.
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import shutil
import subprocess  # nosec
import tempfile
import unicodedata
from pathlib import Path
from typing import Any, Optional, cast

import pandas as pd
from openpyxl import load_workbook
from tabulate import tabulate

from knowledge_flow_backend.core.processors.input.common.base_input_processor import (
    BaseMarkdownProcessor,
    InputConversionError,
)
from knowledge_flow_backend.core.processors.input.excel_processor.excel_extractor import (
    DetectedTable,
    ExcelExtractor,
    SheetSummary,
    _a1,
    _setup_log_file,
    _step,
    log,
)
from knowledge_flow_backend.features.tabular.artifacts import (
    build_table_query_alias,
    build_tabular_table_object_key,
    dataframe_dtype_to_literal,
    utc_now_iso,
)

logger = logging.getLogger(__name__)


def recalc_with_libreoffice(path: str) -> str:
    """
    Open the file with headless LibreOffice, recalculate every formula, and
    export to a temporary xlsx. Return the path of the recalculated xlsx.

    Uses the canonical ``soffice`` binary (the ``libreoffice`` command is only a
    Debian/Ubuntu wrapper around it), matching the shared LibreOffice pattern in
    ``common/legacy_office.py`` and ``pptx_slide_renderer``.
    Requires: sudo apt install libreoffice
    """
    soffice_path = shutil.which("soffice")
    if soffice_path is None:
        raise RuntimeError("LibreOffice executable 'soffice' not found in PATH. Install it: sudo apt install libreoffice")
    outdir = tempfile.mkdtemp()
    # Confine LibreOffice's user profile to the fresh temp dir: the recalc never
    # touches the shared ~/.config/libreoffice profile (no pollution) and concurrent
    # ingestion workers each get an isolated profile instead of racing on one.
    profile_dir = Path(outdir) / ".lo_profile"
    subprocess.run(
        [
            soffice_path,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile_dir.as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            outdir,
            path,
        ],
        check=True,
        capture_output=True,
    )  # nosec: controlled command arguments, shell=False
    # Pick up whatever LibreOffice wrote rather than reconstructing the name (it
    # normalises the filename, e.g. trims trailing spaces). The profile lives in a
    # sub-directory, so the single produced .xlsx directly in outdir is unambiguous.
    produced = [f for f in os.listdir(outdir) if f.lower().endswith(".xlsx")]
    if not produced:
        raise RuntimeError(f"LibreOffice did not produce a file in {outdir}")
    result_path = os.path.join(outdir, produced[0])
    log.info("  [RECALC] LibreOffice → %s", result_path)
    return result_path


class ExcelProcessor(BaseMarkdownProcessor):
    """Excel input processor built around an :class:`ExcelExtractor`.

    For each file, `convert_file_to_markdown` loads the workbook(s) (optionally
    recalculating with headless LibreOffice first), runs the extraction (phases
    A/B) through the extractor, then exports the preview (`output.md`) plus one
    data extract per non-empty table.

    Extraction/export options are class attributes mirroring the original
    workflow driver's constructor defaults; override them on an instance (or
    subclass) to tune the pipeline. No per-file state is kept on the instance —
    the workbook, summaries and output directory flow through method arguments,
    so one processor can safely handle many documents.
    """

    description = "Extracts Excel workbooks into a Markdown summary plus one data extract (CSV or Parquet) per detected table."

    # ------------------------------------------------------------------ #
    # Extraction / export options
    # ------------------------------------------------------------------ #
    include_hidden_sheets: bool = True
    include_hidden_cells: bool = False
    apply_format_masking: bool = True
    keep_split_residuals: bool = True
    split_on_hidden_columns: bool = False
    recalc: bool = True
    coerce_types: bool = True
    keep_headerless_tables: bool = False
    keep_headerless_columns: bool = False
    drop_empty_rows: bool = True
    drop_empty_cols: bool = True
    keep_single_column_tables: bool = False
    # Export: list the column names under each non-empty table of the summary.
    show_column_names: bool = True
    # Export: extract format — "csv" (Excel-friendly) or "parquet" (typed).
    extract_format: str = "parquet"
    # None (or "*") -> process ALL sheets; a name or list restricts the run.
    sheets: list[str] | str | None = None
    debug: bool = False

    # ================================================================== #
    # INPUT-PROCESSOR CONTRACT
    # ================================================================== #

    # 1) Validity — called before anything (process_metadata + process_input)
    def check_file_validity(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in {".xlsx", ".xls", ".xlsm"} and file_path.is_file()

    # 2) Lightweight metadata — "extract_metadata" stage (before Temporal).
    #    Must stay fast: only reads the sheet inventory, never runs extraction.
    def extract_file_metadata(self, file_path: Path) -> dict:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
            try:
                sheet_names = list(wb.sheetnames)
            finally:
                wb.close()
        except Exception as e:
            logger.warning("[PROCESSOR][EXCEL] Failed to read metadata for %s: %s", file_path, e)
            return {}
        return {
            "extras": {
                "excel.sheet_count": len(sheet_names),
                "excel.sheet_names": sheet_names,
            }
        }

    # 3) Core — input stage. Loads the workbook, drives the extractor, and writes
    #    `output.md` (the preview) and `tables.json` under `output_dir`. In
    #    ingestion mode the per-table Parquet extracts are uploaded to the
    #    tabular store instead of being kept under `output_dir` (see `_export`).
    def convert_file_to_markdown(self, file_path: Path, output_dir: Path, document_uid: str | None) -> dict:
        if self.extract_format not in ("csv", "parquet"):
            raise ValueError(f"extract_format must be 'csv' or 'parquet', got {self.extract_format!r}")
        if not file_path.exists():
            raise FileNotFoundError(f"Source file not found: {file_path}")
        output_dir.mkdir(parents=True, exist_ok=True)

        # Route the pipeline log first so the whole run is captured (debug only).
        self._setup_pipeline_log(output_dir)

        try:
            extractor = self._build_extractor(str(file_path))
            summaries = extractor.extract()

            log.info("\n" + "=" * 70)
            log.info("EXPORT")
            log.info("=" * 70)
            self._export(extractor.all_tables, summaries, output_dir, source_path=file_path, document_uid=document_uid)
        except InputConversionError:
            raise
        except Exception as exc:
            raise InputConversionError(f"Excel conversion failed for '{file_path.name}': {exc}") from exc

        return {"doc_dir": str(output_dir), "md_file": str(output_dir / "output.md")}

    # ================================================================== #
    # EXTRACTION — load the workbook(s) and drive the extractor
    # ================================================================== #
    def _setup_pipeline_log(self, output_dir: Path) -> None:
        """Write the pipeline log (`pipeline.log`, next to `output.md`) only in
        debug mode; otherwise no log file is written."""
        if not self.debug:
            return
        output_dir.mkdir(parents=True, exist_ok=True)
        _setup_log_file(str(output_dir / "pipeline.log"))

    def _build_extractor(self, path: str) -> ExcelExtractor:
        """Load the value/formula workbooks (optionally recalculating first with
        headless LibreOffice, then dropping the temporary file) and wire them into
        a fresh :class:`ExcelExtractor` configured from this processor's options."""
        # None (or '*') -> process ALL sheets of the workbook.
        sheets = self.sheets
        if sheets == "*":
            sheets = None
        elif isinstance(sheets, str):
            sheets = [sheets]
        sheets_filter: Optional[list[str]] = sheets

        # Recalculate with LibreOffice, load the workbooks, then drop the
        # temporary file (load_workbook has read it fully into memory).
        # Legacy .xls (binary BIFF) is unreadable by openpyxl: it always goes
        # through the LibreOffice conversion, which outputs an xlsx.
        recalc_dir: Optional[str] = None
        if self.recalc or Path(path).suffix.lower() == ".xls":
            load_path = recalc_with_libreoffice(path)
            recalc_dir = os.path.dirname(load_path)
        else:
            load_path = path
        wb_values = load_workbook(load_path, data_only=True, keep_links=False)
        wb_formulas = load_workbook(load_path, data_only=False, keep_links=False)
        if recalc_dir:
            shutil.rmtree(recalc_dir, ignore_errors=True)
            _step("RECALC", "temporary recalculated file removed")

        return ExcelExtractor(
            wb_values,
            wb_formulas,
            path,
            include_hidden_sheets=self.include_hidden_sheets,
            include_hidden_cells=self.include_hidden_cells,
            apply_format_masking=self.apply_format_masking,
            keep_split_residuals=self.keep_split_residuals,
            split_on_hidden_columns=self.split_on_hidden_columns,
            coerce_types=self.coerce_types,
            keep_headerless_tables=self.keep_headerless_tables,
            keep_headerless_columns=self.keep_headerless_columns,
            drop_empty_rows=self.drop_empty_rows,
            drop_empty_cols=self.drop_empty_cols,
            keep_single_column_tables=self.keep_single_column_tables,
            sheets_filter=sheets_filter,
        )

    # ================================================================== #
    # EXPORT
    # ================================================================== #
    def _export(
        self,
        tables: list[DetectedTable],
        summaries: list[SheetSummary],
        output_dir: Path,
        *,
        source_path: Optional[Path] = None,
        document_uid: Optional[str] = None,
    ) -> None:
        """Export under `output_dir`.

        Always produces, directly under `output_dir`:
          • `tables.json` — the machine-readable catalog of the extracts,
            consumed by the Excel output stage to register each table;
          • `output.md` — the human-readable summary enriched with row counts
            and columns.

        Extract placement depends on the mode:
          • Ingestion mode (`document_uid` and `source_path` provided, extracts
            are Parquet): each extract is uploaded to the tabular store under
            the canonical `tabular/datasets/<uid>/<rev>/` prefix and its catalog
            entry gains `object_key` / `query_alias` — the contract the output
            stage promotes into `metadata.extensions["tabular_multi_v1"]`. The
            LOCAL copies are then removed from `output_dir`: the tabular store
            is the single source of truth, so leaving them would duplicate every
            Parquet table into the document content bucket when
            `save_output(output_dir)` runs. `output.md` still prints each table's
            parquet file NAME as PLAIN TEXT — a reference, not a
            clickable/redirect link. The internal storage prefix
            (`tabular/datasets/<uid>/<rev>/`) is not exposed to the reader.
          • Otherwise (standalone runs, tests, or non-Parquet extracts): the
            extracts stay under `output_dir/<fmt>/` as the only persisted copy
            and `output.md` links to them.
        """
        extract_dir = output_dir / self.extract_format
        extract_dir.mkdir(parents=True, exist_ok=True)

        # 1) one extract + one catalog entry per non-empty table
        entries = self._write_extracts(tables, str(extract_dir))

        # 2) ingestion mode: upload each Parquet extract to its canonical
        #    object key and enrich the catalog entries in place
        if document_uid and source_path is not None:
            self._register_tables(entries, output_dir=output_dir, source_path=source_path, document_uid=document_uid)
        aliases = {entry["table_id"]: entry["query_alias"] for entry in entries if entry.get("query_alias")}
        # Storage location of each registered table, printed as plain text in
        # output.md (the extract lives in the tabular store, a different bucket).
        object_keys = {entry["table_id"]: entry["object_key"] for entry in entries if entry.get("object_key")}

        # 3) once every table lives in the tabular store, drop the local copies
        #    so save_output(output_dir) does not duplicate them into the
        #    document content bucket. Only when ALL entries were registered
        #    (have an object_key); otherwise the local copy is the sole one.
        registered = bool(entries) and all(entry.get("object_key") for entry in entries)
        if registered:
            shutil.rmtree(extract_dir, ignore_errors=True)
            extract_links: Optional[dict[str, str]] = None
            _step("EXPORT", f"{len(entries)} extract(s) in the tabular store; local {extract_dir.name}/ removed (no duplication in output/)")
        else:
            extract_links = {entry["table_id"]: entry["path"] for entry in entries}

        # 4) machine-readable sidecar transported to the output stage
        self._write_tables_sidecar(output_dir / "tables.json", entries)

        # 5) Markdown summary (English)
        md_path = output_dir / "output.md"
        self._write_markdown_summary(str(md_path), extract_links, summaries, aliases=aliases, object_keys=object_keys)
        _step("EXPORT", f"export finished under {output_dir}")

    @staticmethod
    def _table_is_empty(t: DetectedTable) -> bool:
        return t.df is None or t.df.empty or t.etat == "empty"

    @staticmethod
    def _ascii_fold(name: str) -> str:
        """Transliterate accented characters to plain ASCII.

        Diacritics are dropped (``é`` -> ``e``, ``û`` -> ``u``).
        """
        decomposed = unicodedata.normalize("NFKD", name)
        return decomposed.encode("ascii", "ignore").decode("ascii")

    @classmethod
    def _safe_filename(cls, name: str) -> str:
        """Make a table id usable as a file name (spaces, %, ...)."""
        return re.sub(r"[^\w.-]+", "_", cls._ascii_fold(name)).strip("_") or "table"

    @staticmethod
    def _normalize_newlines(df: pd.DataFrame) -> pd.DataFrame:
        """Normalize in-cell line breaks (`\\r\\n`, `\\r`, `\\n`) to a single `\\n`.

        A lone `\\r` would otherwise be read by Excel as a record terminator and
        burst the cell across rows. Combined with a `\\r\\n` line terminator in
        `to_csv`, this keeps multi-line cells intact.
        """
        return cast(pd.DataFrame, df.apply(lambda col: col.map(lambda v: re.sub(r"\r\n?|\n", "\n", v) if isinstance(v, str) else v)))

    def _write_extracts(self, tables: list[DetectedTable], extract_dir: str) -> list[dict]:
        """Write every row of each non-empty table in `extract_format` and
        describe it for the sidecar catalog.

        Returns one catalog entry per non-empty table (see `_table_entry` for the
        shape). An empty table is skipped: no file is written and it produces no
        entry.
        """
        fmt = self.extract_format
        entries: list[dict] = []
        for t in tables:
            df = t.df
            if df is None or self._table_is_empty(t):
                _step("EXPORT", f"{t.id} — empty, no {fmt} written")
                continue
            fname = f"{self._safe_filename(t.id)}.{fmt}"
            dest = os.path.join(extract_dir, fname)
            if fmt == "csv":
                # Fill NaN with "" and normalize in-cell line breaks so Excel
                # keeps multi-line cells intact.
                export = df.astype(object).where(pd.notna(df), "")
                export = self._normalize_newlines(export)
                export.to_csv(
                    dest,
                    index=False,
                    sep=";",
                    encoding="utf-8-sig",
                    lineterminator="\r\n",
                )
            else:
                # Parquet keeps native dtypes, but object columns often mix
                # numbers and text, which the engine can't reduce to one type —
                # cast them to the nullable `string` dtype; typed columns stay.
                export = df.copy()
                obj_cols = [c for c in export.columns if export[c].dtype == object]
                export[obj_cols] = export[obj_cols].astype("string")
                export.to_parquet(dest, index=False)
            rel_path = f"{fmt}/{fname}"
            entries.append(self._table_entry(t, rel_path, df))
            _step("EXPORT", f"{t.id} — {len(df)} row(s) written to {rel_path}")
        return entries

    @staticmethod
    def _table_index(t: DetectedTable) -> int:
        """1-based index of the table within its sheet, parsed from the extractor
        id contract `<sheet>.tN` (A4 numbers tables contiguously per sheet)."""
        try:
            return int(t.id.rsplit(".t", 1)[1])
        except (IndexError, ValueError):
            return 1

    def _table_entry(self, t: DetectedTable, rel_path: str, df: pd.DataFrame) -> dict:
        """Describe one written table for the `tables.json` catalog.

        Columns reuse the platform tabular vocabulary (`dataframe_dtype_to_literal`,
        i.e. the `TabularColumnSchema` shape `{name, dtype}`). Iteration is done by
        position so duplicate column labels are all reported.

        The object-store key, source revision and SQL alias are absent at this
        point: `_register_tables` fills them in when the export runs in
        ingestion mode (document uid available, Parquet extracts).
        """
        return {
            "table_id": t.id,
            "table_index": self._table_index(t),
            "sheet": t.sheet,
            "title": (t.title or "").replace("\n", " ") or None,
            "range": t.plage,
            "data_range": t.plage_donnees,
            "format": self.extract_format,
            "path": rel_path,
            "row_count": int(len(df)),
            "columns": [{"name": str(name), "dtype": dataframe_dtype_to_literal(dtype)} for name, dtype in df.dtypes.items()],
        }

    @staticmethod
    def _compute_source_revision(source_path: Path) -> str:
        """SHA-256 of the source workbook — matches `metadata.file.sha256`
        (hashed from the same input file), so the object keys line up with the
        revision the platform records for the document."""
        hasher = hashlib.sha256()
        with source_path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                hasher.update(chunk)
        return hasher.hexdigest()

    def _register_tables(self, entries: list[dict], *, output_dir: Path, source_path: Path, document_uid: str) -> None:
        """Upload each Parquet extract to its canonical object key and enrich
        its catalog entry in place (ingestion mode only).

        Adds to every entry: `dataset_uid`, `object_key`, `query_alias`,
        `source_revision`, `generated_at`, `file_size_bytes` — exactly the
        fields the output stage needs to build the `tabular_multi_v1`
        extension. Aliases come from the shared deterministic helper and are
        deduplicated within the workbook, so `output.md` and the tabular
        service expose the same relation names.

        Skips silently when there is nothing to register, when extracts are not
        Parquet (SQL mounting requires Parquet), or when no ApplicationContext
        is initialized (standalone/offline runs).
        """
        if not entries:
            return
        if self.extract_format != "parquet":
            logger.warning(
                "[PROCESSOR][EXCEL] extract_format=%r: tables of document %s will not be SQL-registered (Parquet required)",
                self.extract_format,
                document_uid,
            )
            return
        from knowledge_flow_backend.application_context import ApplicationContext

        try:
            context = ApplicationContext.get_instance()
        except RuntimeError:
            logger.warning("[PROCESSOR][EXCEL] No ApplicationContext: skipping table registration for document %s", document_uid)
            return

        content_store = context.get_content_store()
        artifacts_prefix = context.get_config().storage.tabular_store.artifacts_prefix
        source_revision = self._compute_source_revision(source_path)
        generated_at = utc_now_iso()

        used_aliases: set[str] = set()
        for entry in entries:
            base_alias = build_table_query_alias(document_uid, entry["sheet"], entry["table_index"])
            query_alias = base_alias
            suffix = 2
            while query_alias in used_aliases:
                query_alias = f"{base_alias}_{suffix}"
                suffix += 1
            used_aliases.add(query_alias)

            local_file = output_dir / entry["path"]
            object_key = build_tabular_table_object_key(
                artifacts_prefix=artifacts_prefix,
                document_uid=document_uid,
                source_revision=source_revision,
                table_file_name=local_file.name,
            )
            stored = content_store.put_file(object_key, local_file, content_type="application/vnd.apache.parquet")
            entry.update(
                {
                    "dataset_uid": document_uid,
                    "object_key": stored.key,
                    "query_alias": query_alias,
                    "source_revision": source_revision,
                    "generated_at": generated_at,
                    "file_size_bytes": stored.size,
                }
            )
            _step("EXPORT", f'{entry["table_id"]} — registered as "{query_alias}" at {stored.key}')

    @staticmethod
    def _write_tables_sidecar(sidecar_path: Path, entries: list[dict]) -> None:
        """Write `tables.json` — a JSON array with one object per non-empty table
        (`table_id`, `sheet`, `title`, `range`, `data_range`, `format`, `path`,
        `row_count`, `columns`). It is the machine-readable counterpart of
        `output.md`, transported to the output stage which promotes each table
        into the tabular store.
        """
        with open(sidecar_path, "w", encoding="utf-8") as f:
            json.dump(entries, f, ensure_ascii=False, indent=2)
        _step("EXPORT", f"sidecar written to {sidecar_path} ({len(entries)} table(s))")

    def _write_markdown_summary(
        self,
        md_path: str,
        csv_links: Optional[dict[str, str]],
        summaries: list[SheetSummary],
        aliases: Optional[dict[str, str]] = None,
        object_keys: Optional[dict[str, str]] = None,
    ) -> None:
        """Write `output.md` — the human preview AND the SQL catalog contract.

        `csv_links` maps each table id to its local extract path; pass `None` in
        ingestion mode, where the extracts live in the tabular store (not next
        to `output.md`) so no local link is emitted.

        `object_keys` maps each table id to its storage key in the tabular store;
        only the file NAME (basename) is printed, as plain text (a reference, not
        a redirect link). The internal storage prefix
        (`tabular/datasets/<uid>/<rev>/`) is plumbing that does not concern the
        reader, so it is stripped from the summary.

        In ingestion mode `aliases` maps each table id to its exact
        `query_alias`: the relation name the tabular service mounts in DuckDB.
        Downstream agents read this file to know which tables exist and how to
        address them in SQL, so the alias printed here must never diverge from
        the one stored in `tables.json` / `tabular_multi_v1`.
        """
        lines = ["# Extraction summary", ""]
        for s in summaries:
            lines += self._md_sheet_lines(s, csv_links, debug=self.debug, show_column_names=self.show_column_names, aliases=aliases, object_keys=object_keys)
        if self.debug:
            lines += self._md_debug_preview_lines(summaries)
        with open(md_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines).rstrip() + "\n")
        _step("EXPORT", f"summary written to {md_path}")

    def _md_debug_preview_lines(self, summaries: list[SheetSummary]) -> list[str]:
        """Debug section: first 10 rows of each table, in Markdown."""
        lines = ["", "---", "", "## Debug: table previews (first 10 rows)", ""]
        for s in summaries:
            for t in s.tables:
                if self._table_is_empty(t):
                    continue
                lines += [f"### {t.id}", ""]
                preview = t.df.head(10).astype(object).where(pd.notna(t.df.head(10)), "")
                lines += tabulate(preview, headers="keys", tablefmt="pipe", showindex=False).splitlines()
                lines.append("")
        return lines

    @staticmethod
    def _residual_value(content: Any) -> str:
        """All the non-empty values of a residual.

        A single value stays on the line (`value="OK"`); several values (stacked
        cells, or a single multi-line cell) are unrolled below it, one per line.
        """
        if content is None:
            return ""
        # content is usually the block's 2D grid, but a string or flat list also
        # works: flatten everything, then split each value on its line breaks.
        items = content if isinstance(content, list) else [content]
        flat: list[str] = []
        for item in items:
            cells = item if isinstance(item, list) else [item]
            for v in cells:
                if v is None:
                    continue
                for part in str(v).split("\n"):
                    if part.strip():
                        flat.append(part.strip())
        if not flat:
            return ""
        if len(flat) == 1:
            return f'  value="{flat[0]}"'
        body = "\n".join(flat)
        return f'  value="\n{body}\n"'

    @staticmethod
    def _md_sheet_lines(
        s: SheetSummary,
        csv_links: Optional[dict[str, str]] = None,
        debug: bool = False,
        show_column_names: bool = False,
        aliases: Optional[dict[str, str]] = None,
        object_keys: Optional[dict[str, str]] = None,
    ) -> list[str]:
        vis = "hidden" if not s.visible else "visible"
        lines = [f"## Sheet: {s.name}  ({vis}, coverage={s.coverage * 100:.0f}%)", ""]
        if s.tables:
            lines.append(f"Tables ({len(s.tables)}):")
            for t in s.tables:
                lines += ExcelProcessor._md_table_lines(t, csv_links, debug=debug, show_column_names=show_column_names, aliases=aliases, object_keys=object_keys)
            lines.append("")
        if s.residuals:
            lines.append(f"Unextracted residuals ({len(s.residuals)}):")
            for r in s.residuals:
                val_str = ExcelProcessor._residual_value(r.content)
                lines.append(f'- range="{_a1(*r.bbox)}"  type={r.type}{val_str}')
            lines.append("")
        return lines

    @staticmethod
    def _md_table_lines(
        t: DetectedTable,
        csv_links: Optional[dict[str, str]] = None,
        debug: bool = False,
        show_column_names: bool = False,
        aliases: Optional[dict[str, str]] = None,
        object_keys: Optional[dict[str, str]] = None,
    ) -> list[str]:
        n_rows = len(t.df) if t.df is not None else 0
        title = (t.title or "").replace("\n", " ")
        context = " | ".join(c.replace("\n", " ") for c in t.context) if t.context else ""
        range_part = f"  range={t.plage}  data_range={t.plage_donnees or '-'}"
        alias = (aliases or {}).get(t.id)
        alias_part = f'  query_alias="{alias}"' if alias else ""
        # Parquet extract file NAME only, as plain text (not a markdown link).
        # The full storage key (tabular/datasets/<uid>/<rev>/…) is internal
        # plumbing that does not concern the reader — only the file name is
        # shown. Tables are addressed in SQL via query_alias, never this path.
        object_key = (object_keys or {}).get(t.id)
        object_key_part = f'  parquet="{os.path.basename(object_key)}"' if object_key else ""
        line = f'- "{t.id}"{alias_part}{object_key_part}{range_part}  state={t.etat}  rows={n_rows}  title="{title}"  context="{context}"  '
        if csv_links is not None:
            link = csv_links.get(t.id)
            label = os.path.splitext(link)[1].lstrip(".").upper() if link else ""
            line += f" — [{label}]({link})" if link else " — empty"
        lines = [line]
        # Column-names sub-bullet, only for NON-empty tables.
        if show_column_names and t.df is not None and not ExcelProcessor._table_is_empty(t):
            cols = " | ".join(f'"{c}"' for c in t.df.columns)
            lines.append(f"    column_name: {cols}")
        return lines
