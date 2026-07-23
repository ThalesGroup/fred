"""
excel_extractor.py
==================
Excel extraction — implementation of the processing pipeline specification.

This module holds the processing RULES: the ``ExcelExtractor`` class receives
the already-loaded workbooks and applies every step of the specification, named
after its code (a1_ ... a5_, b1_ ... b5_). Its ``extract()`` orchestrator chains
phases A and B in the exact order of the spec and returns the per-sheet
summaries and the flat list of detected tables. It performs no file I/O.

The data structures that flow between steps and the pure helpers also live here,
at module level. The INFRASTRUCTURE (file handling, LibreOffice recalculation,
workflow driver and output generation) lives in ``excel_processor.py``
(``ExcelProcessor``); the test-workbook builder in ``build_test_excel.py``.

Mapping to the specification
----------------------------
  Principles     -> permissive extraction then typing; capture before pandas destroys
  A1 inventory   -> inventory + enriched summary (tables + residuals per sheet)
  A2 load        -> data_only read, capture merges / outline / errors
  A3 detect      -> connected components (blocks separated by empty rows/columns)
  A4 split       -> split stacked tables
                    (full-width merged row = title + context)
  A5 strip_cols  -> remove merged label columns at the start of a table
                    (vertical merge >=80% of the height -> value moved to context)
  B1 orientation -> transposed / cross-tab / normal
  B2 to_dataframe-> auto-fill merged cells + pandas conversion (first row = header)
  B3 check_empty -> table state: "empty" / "non-empty"
  B4 clean       -> column recognition (dates, numeric: strip thousands spaces, text)
  B5 validate    -> provenance attached
  Cross-cutting  -> LLM fallback + observability (logs at each step)

Dependencies: numpy, pandas, openpyxl (required); dlt (optional).
"""

from __future__ import annotations

import datetime as dt
import logging
import math
import re
from dataclasses import dataclass, field
from typing import Any, Optional, cast

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

# Dedicated pipeline logger. It stays independent of the host application's
# logging: no `logging.basicConfig()` (that belongs to the app, not to a library
# module) and `propagate = False` so pipeline messages never leak into the app's
# handlers. The only output is the optional file handler wired up in debug mode
# (see `_setup_log_file`).
log = logging.getLogger("excel_pipeline")
log.setLevel(logging.INFO)
log.propagate = False

_log_file_handler: Optional[logging.FileHandler] = None


def _setup_log_file(path: str) -> None:
    global _log_file_handler
    if _log_file_handler:
        log.removeHandler(_log_file_handler)
        _log_file_handler.close()
    _log_file_handler = logging.FileHandler(path, mode="w", encoding="utf-8")
    _log_file_handler.setFormatter(logging.Formatter("%(message)s"))
    log.addHandler(_log_file_handler)


def _step(code: str, msg: str) -> None:
    """Log one step of the pipeline, prefixed by its spec code (A1, B2, ...)."""
    log.info("  [%s] %s", code, msg)


# --------------------------------------------------------------------------- #
# Data structures that flow between steps
# --------------------------------------------------------------------------- #
@dataclass
class DetectedTable:
    """A detected table, enriched as it passes through the B steps."""

    id: str
    sheet: str
    bbox: tuple[int, int, int, int]  # (r0, c0, r1, c1) 0-indexed inclusive — full footprint (title + context + width)
    grid: np.ndarray  # raw slice (dtype object)
    local_merges: list = field(default_factory=list)  # merges in local coords
    title: Optional[str] = None
    context: list[str] = field(default_factory=list)  # merged rows below the title (subtitles, notes)
    data_bbox: Optional[tuple[int, int, int, int]] = None  # KEPT range: body only (excludes title/context and dropped rows/cols)
    orientation: str = "normal"
    df: Optional[pd.DataFrame] = None
    has_column_names: bool = True  # False if the header was fully auto-generated (set in B2)
    status: str = "ok"
    etat: str = "non-empty"  # "empty" / "non-empty" (set in B3)
    provenance: dict = field(default_factory=dict)

    @property
    def plage(self) -> str:
        """A1 notation of the full footprint (title + context + original width)."""
        return _a1(*self.bbox)

    @property
    def plage_donnees(self) -> Optional[str]:
        """A1 notation of the KEPT range (body only, excluding title/context and
        dropped rows/columns), or None if splitting never filled it in."""
        return _a1(*self.data_bbox) if self.data_bbox else None


@dataclass
class Residual:
    """An unextracted block: free title, note, isolated cells, or parsing failure."""

    sheet: str
    bbox: tuple[int, int, int, int]
    type: str
    status: str = "residual"
    content: Any = None


@dataclass
class CandidateBlock:
    """Candidate block from A3 (a filtered connected component), to be re-split in A4."""

    sheet: str
    row_abs: list[int]  # absolute sheet row for each row of the filtered block
    col_abs: list[int]  # absolute sheet column for each column of the filtered block
    c0: int  # first ORIGINAL column of the block (before filtering/segmentation)
    c1: int  # last original column
    grid: np.ndarray
    local_merges: list = field(default_factory=list)


@dataclass
class SheetSummary:
    name: str
    visible: bool
    used_range: str
    n_merges: int
    has_formulas: bool
    tables: list = field(default_factory=list)  # list[DetectedTable]
    residuals: list = field(default_factory=list)  # list[Residual]
    coverage: float = 0.0


# --------------------------------------------------------------------------- #
# General helpers
# --------------------------------------------------------------------------- #
def _nonempty(v: Any) -> bool:
    return v is not None and not (isinstance(v, float) and math.isnan(v)) and str(v).strip() != ""


def _is_number(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool) and not (isinstance(v, float) and math.isnan(v))


def _nonempty_mask(arr: np.ndarray) -> np.ndarray:
    """Boolean mask of the non-empty cells of a grid (empty-safe)."""
    return np.vectorize(_nonempty)(arr) if arr.size else np.zeros(arr.shape, bool)


def _count_nonempty(arr: np.ndarray) -> int:
    """Number of non-empty cells in a grid (empty-safe)."""
    return int(_nonempty_mask(arr).sum())


# A block is a table area (vs. a residual) only if it is at least 2×2 and holds
# at least 4 non-empty cells. Same rule guards A3's candidates and A4's split
# bodies, so it lives in one place.
def _is_table_area(height: int, width: int, n_cells: int) -> bool:
    return height >= 2 and width >= 2 and n_cells >= 4


# Excel error literals all start with "#" and usually end with "!" or "?"
# (#VALUE!, #REF!, #DIV/0!, ...); #N/A and #GETTING_DATA are the exceptions the
# trailing-punctuation rule would miss.
_EXCEL_ERROR_RE = re.compile(r"^#(N/A|GETTING_DATA|.*[!?])$")


def _is_excel_error(v: Any) -> bool:
    """True for any Excel error value (#N/A, #VALUE!, #REF!, ...).

    openpyxl returns the cached error as a string with data_only=True; such a
    cell carries no real value, so the pipeline treats it as empty."""
    return isinstance(v, str) and bool(_EXCEL_ERROR_RE.match(v.strip()))


def _is_data_like(v: Any) -> bool:
    """A "data" value = a number or a date (as opposed to a header label)."""
    return _is_number(v) or isinstance(v, (dt.date, dt.datetime))


def _format_hides_value(value: Any, number_format: Optional[str]) -> bool:
    """Return True if Excel's number format renders this value INVISIBLE on screen.

    A format can have up to 4 sections separated by `;` (positive ; negative ;
    zero ; text). If the section that applies to the value is empty, Excel shows
    nothing (the cell looks empty though it holds a cached value) — e.g. `'0;\\-0;;'`
    hides zeros. Only numbers are handled.
    """
    if not _is_number(value) or not number_format:
        return False
    # A naive split on `;` is enough here: escaped/quoted `;` don't appear in the
    # usual hiding formats.
    sections = number_format.split(";")
    n = len(sections)
    if n == 1:
        idx = 0
    elif n == 2:
        idx = 0 if value >= 0 else 1  # zero follows the positive section
    else:
        idx = 0 if value > 0 else 1 if value < 0 else 2
    return sections[idx].strip() == ""


def _a1(r0: int, c0: int, r1: int, c1: int) -> str:
    """0-indexed coords -> Excel notation 'A1:E5'."""
    return f"{get_column_letter(c0 + 1)}{r0 + 1}:{get_column_letter(c1 + 1)}{r1 + 1}"


def _hidden_cols(ws) -> set[int]:
    """0-based indices of the hidden columns.

    openpyxl groups contiguous same-style columns into a single `ColumnDimension`
    with `.min`/`.max` (a hidden range A->G is one entry min=1, max=7). We expand
    each entry over its min..max interval; reading column by column would miss the
    middle of a range.
    """
    hidden: set[int] = set()
    for dim in ws.column_dimensions.values():
        if dim.hidden:
            hidden.update(range(dim.min - 1, dim.max))
    return hidden


def _real_extent(ws) -> tuple[int, int]:
    """Real (nrows, ncols) from the cells that actually hold content.

    max_row/max_column follow the <dimension> tag, which can be inflated by
    deleted data/styles or LibreOffice artefacts. Scanning ws._cells for values
    that are neither None nor 0 keeps the grid bounded to real content; falls back
    to max_row/max_column if ws._cells is unavailable.
    """
    cells = getattr(ws, "_cells", None)
    if not cells:
        return (0, 0) if cells is not None else (ws.max_row, ws.max_column)
    valued = [(r, c) for (r, c), cell in cells.items() if cell.value is not None and cell.value != 0]
    if not valued:
        return 0, 0
    max_row = max(r for r, _ in valued)
    max_col = max(c for _, c in valued)
    return max_row, max_col


# Numeric handling: a numeric column is RECOGNISED but NOT converted — only the
# thousands-grouping whitespace is stripped (currency, %, decimal comma kept).
_THOUSAND_SEP = re.compile(r"[\s   ]")
_CURRENCY = re.compile(r"[€$£¥]")
# A whitespace sitting *between two digits* is a thousands separator.
_THOUSAND_BETWEEN = re.compile(r"(?<=\d)[\s   ](?=\d)")
_NUM_SENTINELS = {"", "-", "–", "N/A", "NA", "n/a", "TBD"}


def _despace_thousands(s: Any) -> str:
    """Strip the whitespace grouping thousands (between two digits) and trim.

    Currency symbols, percent signs and decimal commas are left untouched:
    "1 234,56 €" -> "1234,56 €", "15 %" -> "15 %", "1 200" -> "1200"."""
    return _THOUSAND_BETWEEN.sub("", str(s).strip())


def _looks_numeric(v: Any) -> bool:
    """True if the token reads as a number once currency/%/locale are ignored.

    Used only to RECOGNISE a numeric column; the value itself is never coerced."""
    raw = str(v).strip().rstrip("%").strip()
    raw = _CURRENCY.sub("", raw)
    raw = _THOUSAND_SEP.sub("", raw)
    raw = raw.replace(",", ".")  # FR locale: decimal comma
    if raw in _NUM_SENTINELS:
        return False
    try:
        float(raw)
        return True
    except ValueError:
        return False


class ExcelExtractor:
    """Processing rules of the specification (phases A and B).

    Operates on already-loaded workbooks (`wb_values`, `wb_formulas`) — the file
    handling and recalculation are `ExcelProcessor`'s job. Each step is a method
    named after its spec code (a1_ ... a5_, b1_ ... b5_); `extract()` chains them
    in the exact order of the specification, logging as it goes, and returns the
    per-sheet summaries. The flat list of the tables kept is exposed as
    `all_tables` for the export phase.
    """

    # A4 — minimum fraction of the block width that a merged row must cover to be
    # treated as a table separator (title/context).
    SPLIT_SEPARATOR_WIDTH_RATIO: float = 0.8

    # A5 — minimum fraction of the body height that a vertical merge must cover for
    # a leading column to be treated as a label (moved to the context and then
    # removed from the table).
    STRIP_LABEL_COLUMN_HEIGHT_RATIO: float = 0.8

    # B1 — orientation-detection thresholds (transposed / cross-tab / normal):
    #   ORIENT_HEADER_NUMERIC_RATIO: fraction of the first row that must be
    #       "data-like" (number/date) to suspect a transpose/cross-tab.
    #   ORIENT_FIRST_COL_TEXT_RATIO: fraction of the first column that must be
    #       textual for the same suspicion.
    #   ORIENT_INNER_NUMERIC_RATIO: fraction of the inner body that must be
    #       data-like to confirm a genuine transpose.
    ORIENT_HEADER_NUMERIC_RATIO: float = 0.6
    ORIENT_FIRST_COL_TEXT_RATIO: float = 0.6
    ORIENT_INNER_NUMERIC_RATIO: float = 0.8

    def __init__(
        self,
        wb_values,
        wb_formulas,
        path: str,
        include_hidden_sheets: bool = True,
        include_hidden_cells: bool = True,
        apply_format_masking: bool = True,
        keep_split_residuals: bool = False,
        split_on_hidden_columns: bool = False,
        coerce_types: bool = True,
        keep_headerless_tables: bool = True,
        keep_headerless_columns: bool = True,
        drop_empty_rows: bool = True,
        drop_empty_cols: bool = True,
        keep_single_column_tables: bool = True,
        sheets_filter: Optional[list[str]] = None,
    ):
        self.wb_values = wb_values
        self.wb_formulas = wb_formulas
        # Source path — used only to stamp each table's provenance.
        self.path = path
        self.include_hidden_sheets = include_hidden_sheets
        self.include_hidden_cells = include_hidden_cells
        # A2: neutralize values that Excel hides via their number format (True ->
        # faithful to the display; False -> keep raw cached values, faster read).
        self.apply_format_masking = apply_format_masking
        # Keep the residuals produced by SPLITTING a table: dropped column
        # segments (A3), too-thin body / orphan title (A4).
        self.keep_split_residuals = keep_split_residuals
        # A3: a hidden column in the MIDDLE of a block splits it into runs. True ->
        # keep only the widest run; False (default) -> keep all columns together.
        self.split_on_hidden_columns = split_on_hidden_columns
        # B4: type the columns (num/date/bool/id) or keep the raw data.
        self.coerce_types = coerce_types
        # B2: keep (True) or drop (False) tables whose first row supplies no column
        # name (header fully auto-generated, "col_0...").
        self.keep_headerless_tables = keep_headerless_tables
        # B2: keep (True) or drop (False) individual columns whose first-row cell
        # supplies no title ("col_N"). Independent of keep_headerless_tables.
        self.keep_headerless_columns = keep_headerless_columns
        # B3: removal of fully empty rows / columns.
        self.drop_empty_rows = drop_empty_rows
        self.drop_empty_cols = drop_empty_cols
        # B3 (option): if after cleanup a table is reduced to a single column,
        # keep it (True) or drop it (False).
        self.keep_single_column_tables = keep_single_column_tables
        # None -> process ALL sheets of the workbook.
        self.sheets_filter = sheets_filter
        self.summaries: list[SheetSummary] = []
        self.all_tables: list[DetectedTable] = []

    # ================================================================== #
    # ORCHESTRATOR — the sequence from the spec, visible at a glance
    # ================================================================== #
    def extract(self) -> list[SheetSummary]:
        """Run phases A and B; populate `summaries` and `all_tables`, and return
        the summaries. No file is written here (that is phase C, in the processor)."""
        log.info("=" * 70)
        log.info("PHASE A — document level")
        log.info("=" * 70)
        self.a1_inventory()

        self.all_tables = []
        for summary in self.summaries:
            if not summary.visible and not self.include_hidden_sheets:
                log.info('\n── sheet "%s" skipped (hidden)', summary.name)
                continue
            log.info('\n── sheet "%s" %s', summary.name, "(hidden)" if not summary.visible else "")
            grid, structure = self.a2_load_and_capture(summary.name)
            candidates, residuals = self.a3_detect_tables(summary.name, grid, structure)
            tables, split_residuals = self.a4_split_stacked_tables(candidates)
            self.a5_strip_leading_label_columns(tables)
            residuals += split_residuals
            summary.tables, summary.residuals = tables, residuals
            summary.coverage = self._coverage(grid, tables + [], residuals)

            log.info("  PHASE B — %d table(s) to process", len(tables))
            kept_tables: list[DetectedTable] = []
            for table in tables:
                try:
                    self.b1_detect_orientation(table)
                    self.b2_to_dataframe(table)
                    # B2 (option): drop a table that has no column names.
                    if not self.keep_headerless_tables and not table.has_column_names:
                        table.status = "dropped_headerless"
                        _step("B2", f"{table.id} — no column names, dropped")
                        summary.residuals.append(Residual(summary.name, table.bbox, "headerless_table", status="dropped", content="table without column names"))
                        continue
                    self.b3_check_empty(table)
                    # B3 (option): drop a table reduced to a single column.
                    if not self.keep_single_column_tables and table.df is not None and table.df.shape[1] <= 1:
                        table.status = "dropped_single_column"
                        _step("B3", f"{table.id} — reduced to a single column, dropped")
                        summary.residuals.append(Residual(summary.name, table.bbox, "single_column_table", status="dropped", content="table reduced to a single column"))
                        continue
                    self.b4_clean_and_coerce(table)
                    self.b5_validate_and_tag(table)
                    kept_tables.append(table)
                    self.all_tables.append(table)
                except Exception as exc:  # last-resort fallback
                    table.status = "failed"
                    kept_tables.append(table)  # kept in the summary (status "failed")
                    summary.residuals.append(Residual(summary.name, table.bbox, "failed_table", status="failed", content=str(exc)))
            summary.tables = kept_tables
        return self.summaries

    # ================================================================== #
    # PHASE A
    # ================================================================== #
    def a1_inventory(self) -> list[SheetSummary]:
        """A1 — triage map of the workbook (tables/residuals are added in A3+)."""
        all_names = self.wb_values.sheetnames
        if self.sheets_filter is not None:
            unknown = [n for n in self.sheets_filter if n not in all_names]
            if unknown:
                raise ValueError(f"Sheet(s) not found: {unknown}. Available: {all_names}")
            names_to_process = [n for n in all_names if n in self.sheets_filter]
        else:
            names_to_process = all_names

        for name in names_to_process:
            ws_v = self.wb_values[name]
            ws_f = self.wb_formulas[name]
            has_formulas = any(isinstance(c.value, str) and c.value.startswith("=") for row in ws_f.iter_rows() for c in row)
            s = SheetSummary(
                name=name,
                visible=(ws_v.sheet_state == "visible"),
                used_range=ws_v.dimensions,
                n_merges=len(ws_v.merged_cells.ranges),
                has_formulas=has_formulas,
            )
            self.summaries.append(s)
            _step("A1", f'"{name}" : {s.used_range}, {s.n_merges} merge(s), formulas={has_formulas}, {"visible" if s.visible else "HIDDEN"}')
        return self.summaries

    def a2_load_and_capture(self, sheet_name: str):
        """A2 — frozen values (data_only) + capture of everything pandas would erase."""
        ws = self.wb_values[sheet_name]
        # Real bounds, not the (possibly inflated) <dimension> tag, so the
        # np.empty grid stays sized to the actual content. See _real_extent.
        nrows, ncols = _real_extent(ws)
        grid = np.empty((nrows, ncols), dtype=object)
        if not self.apply_format_masking:
            # Fast path: raw values, without reading each cell's number_format.
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols, values_only=True)):
                for c, v in enumerate(row):
                    grid[r, c] = v
        else:
            masked_display = 0
            zeros_hidden = 0
            # When the sheet's "show a zero in cells that have a zero value"
            # setting is off, Excel hides ALL numeric zeros regardless of format.
            # We treat them as empty to stay faithful to the display. showZeros is
            # False only when off.
            hide_zeros = ws.sheet_view.showZeros is False
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols)):
                for c, cell in enumerate(row):
                    v = cell.value
                    # A value hidden by its number format is invisible on screen
                    # -> treated as empty.
                    if v is not None and _format_hides_value(v, cell.number_format):
                        masked_display += 1
                        v = None
                    elif hide_zeros and _is_number(v) and v == 0:
                        zeros_hidden += 1
                        v = None
                    grid[r, c] = v
            if masked_display:
                _step("A2", f"[FORMAT] {masked_display} cell(s) hidden by their Excel format -> treated as empty")
            if zeros_hidden:
                _step("A2", f'[ZEROS] {zeros_hidden} zero(s) hidden by the sheet "show a zero" setting (showZeros=False) -> treated as empty')

        # Neutralize Excel error values (#N/A, #REF!, ...): they carry no real
        # value, so record their positions and blank them out before detection.
        error_cells: set[tuple[int, int]] = set()
        if grid.size:
            err_mask = np.vectorize(_is_excel_error)(grid)
            if err_mask.any():
                error_cells = {(int(r), int(c)) for r, c in zip(*np.where(err_mask))}
                grid[err_mask] = None
                _step("A2", f"[ERROR] {len(error_cells)} Excel error cell(s) (#N/A, #REF!, ...) -> treated as empty")

        # Detect hidden rows/columns but keep the grid intact (so component
        # detection isn't fragmented); filtering happens at slicing time in A3.
        hidden_rows: set = set()
        hidden_cols: set = set()
        if not self.include_hidden_cells:
            hidden_rows = {i - 1 for i, d in ws.row_dimensions.items() if d.hidden}
            hidden_cols = _hidden_cols(ws)
            hidden_col_letters = sorted(get_column_letter(c + 1) for c in hidden_cols)
            hidden_row_nums = sorted(i for i, d in ws.row_dimensions.items() if d.hidden)
            _step("A2", f"[HIDDEN] Excel rows={hidden_row_nums}  Excel columns={hidden_col_letters}  (0-based: rows={sorted(hidden_rows)}, cols={sorted(hidden_cols)})")

        # Merges -> 0-indexed inclusive coords
        merges = []
        for rng in ws.merged_cells.ranges:
            merges.append((rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1))

        # Outline levels (native Excel subtotals). Error cells were captured and
        # blanked above (their positions are in `error_cells`).
        outline = {i - 1: d.outline_level for i, d in ws.row_dimensions.items()}

        structure = {"merges": merges, "outline": outline, "errors": error_cells, "hidden_rows": hidden_rows, "hidden_cols": hidden_cols}
        _step("A2", f"grid {nrows}×{ncols} read, {len(merges)} merge(s) captured, {len(error_cells)} error cell(s)")
        if merges:
            _step("A2", f"[MERGES] {', '.join(_a1(*m) for m in merges)}")
        return grid, structure

    def a3_detect_tables(self, sheet_name: str, grid: np.ndarray, structure: dict):
        """A3 — connected components: blocks separated by empty rows/columns.

        Returns (candidates, residuals). A block big enough to be a table area
        becomes a candidate (with hidden rows/columns filtered out) that A4 then
        re-splits into tables; anything smaller is classified as a residual.
        """
        mask = _nonempty_mask(grid)
        comps = self._drop_contained_components(self._connected_components(mask))

        candidates, residuals = [], []
        for bbox in comps:
            r0, c0, r1, c1 = bbox
            block = grid[r0 : r1 + 1, c0 : c1 + 1]
            n_cells = int(mask[r0 : r1 + 1, c0 : c1 + 1].sum())
            if _is_table_area(r1 - r0 + 1, c1 - c0 + 1, n_cells):
                candidate, dropped = self._build_candidate_block(sheet_name, grid, structure, bbox, block)
                candidates.append(candidate)
                residuals.extend(dropped)
            else:
                residual = self._classify_residual(sheet_name, structure, bbox, block, n_cells)
                if residual is not None:
                    residuals.append(residual)
        return candidates, residuals

    @staticmethod
    def _drop_contained_components(comps: list) -> list:
        """Drop any component whose bounding box is fully contained in another's:
        they create misleading geometric overlaps (the cells are disjoint, but the
        bounding rectangles overlap)."""

        def contained(inner, outer):
            return inner != outer and outer[0] <= inner[0] and outer[1] <= inner[1] and outer[2] >= inner[2] and outer[3] >= inner[3]

        return [c for c in comps if not any(contained(c, other) for other in comps)]

    def _build_candidate_block(self, sheet_name: str, grid: np.ndarray, structure: dict, bbox: tuple, block: np.ndarray):
        """Turn a table-area component into a `CandidateBlock` for A4.

        Physically removes hidden rows/columns from the block, optionally segments
        it on a hidden middle column (`split_on_hidden_columns`), and remaps the
        merges into the reduced block's coordinates. Returns (candidate,
        residuals) — the residuals being the dropped column segments, if any.
        """
        r0, c0, r1, c1 = bbox
        residuals: list = []
        blk_label = f"{sheet_name}@{_a1(r0, c0, r1, c1)}"

        hidden_rows = structure.get("hidden_rows", set())
        hidden_cols = structure.get("hidden_cols", set())
        vis_r = [i for i, r in enumerate(range(r0, r1 + 1)) if r not in hidden_rows]
        vis_c = [i for i, c in enumerate(range(c0, c1 + 1)) if c not in hidden_cols]
        masked_rows = [r0 + i for i in range(r1 - r0 + 1) if i not in vis_r]
        masked_cols = [c0 + i for i in range(c1 - c0 + 1) if i not in vis_c]
        if masked_rows or masked_cols:
            _step("A3", f"[FILTER] {blk_label} : abs rows={masked_rows} abs cols={masked_cols} -> removed from the block")
        else:
            _step("A3", f"[FILTER] {blk_label} : no hidden row/column in this block")

        if self.split_on_hidden_columns:
            vis_c = self._segment_on_hidden_columns(sheet_name, grid, bbox, blk_label, vis_c, residuals)

        if len(vis_r) < block.shape[0] or len(vis_c) < block.shape[1]:
            block = block[np.ix_(vis_r, vis_c)]
        local_merges = self._remap_merges(structure["merges"], bbox, vis_r, vis_c)

        # row_abs[i]/col_abs[j] = absolute sheet row/column of cell (i, j) in the
        # filtered block (after removing hidden rows/columns and dropped segments).
        row_abs = [r0 + i for i in vis_r]
        col_abs = [c0 + j for j in vis_c]
        candidate = CandidateBlock(sheet_name, row_abs, col_abs, c0, c1, block.copy(), local_merges)
        _step("A3", f"candidate block {blk_label} ({len(row_abs)}×{len(vis_c)}) -> A4")
        return candidate, residuals

    @staticmethod
    def _remap_merges(merges: list, bbox: tuple, vis_r: list[int], vis_c: list[int]) -> list:
        """Remap the sheet merges into the filtered block's coordinates.

        Each merge is CLIPPED to the kept rows/columns (rather than dropped when a
        corner falls on a removed row/column): a full-width title merge thus stays
        full-width in the reduced block and A4 detects it. Merges lying entirely
        outside the block, or with no kept row/column left, are skipped.
        """
        r0, c0, r1, c1 = bbox
        rmap = {old: new for new, old in enumerate(vis_r)}
        cmap = {old: new for new, old in enumerate(vis_c)}
        remapped = []
        for mr0, mc0, mr1, mc1 in merges:
            if not (mr0 >= r0 and mc0 >= c0 and mr1 <= r1 and mc1 <= c1):
                continue
            a, b, cc, d = mr0 - r0, mc0 - c0, mr1 - r0, mc1 - c0  # block-local coords
            rk = [k for k in vis_r if a <= k <= cc]
            ck = [k for k in vis_c if b <= k <= d]
            if not rk or not ck:  # merge outside the kept area
                continue
            remapped.append((rmap[rk[0]], cmap[ck[0]], rmap[rk[-1]], cmap[ck[-1]]))
        return remapped

    def _segment_on_hidden_columns(self, sheet_name: str, grid: np.ndarray, bbox: tuple, blk_label: str, vis_c: list[int], residuals: list) -> list[int]:
        """A3 (`split_on_hidden_columns`) — a hidden column IN THE MIDDLE cuts the
        visible columns into several contiguous runs. Keep only the widest run
        (ties -> the rightmost); the others become residuals (when kept). Returns
        the kept run of local column indices, unchanged when there is a single run.
        """
        r0, c0, r1, _ = bbox
        runs: list[list[int]] = []
        for ci in vis_c:
            if runs and ci == runs[-1][-1] + 1:
                runs[-1].append(ci)
            else:
                runs.append([ci])
        if len(runs) <= 1:
            return vis_c
        kept = max(runs, key=lambda run: (len(run), run[0]))
        for run in runs:
            if run is kept:
                continue
            rc0, rc1 = c0 + run[0], c0 + run[-1]
            if self.keep_split_residuals:
                residuals.append(Residual(sheet_name, (r0, rc0, r1, rc1), "off_segment_columns", content=grid[r0 : r1 + 1, rc0 : rc1 + 1].tolist()))
            dest = "-> residual" if self.keep_split_residuals else "(not kept)"
            _step("A3", f"[FILTER] {blk_label} : column segment {_a1(r0, rc0, r1, rc1)} ({len(run)} col.) dropped {dest}")
        _step("A3", f"[FILTER] {blk_label} : kept segment {_a1(r0, c0 + kept[0], r1, c0 + kept[-1])} ({len(kept)} col.)")
        return kept

    def _classify_residual(self, sheet_name: str, structure: dict, bbox: tuple, block: np.ndarray, n_cells: int):
        """Classify a non-table component as a `Residual`, or return None when it
        is fully hidden (all its rows OR all its columns hidden) and skipped."""
        r0, c0, r1, c1 = bbox
        hidden_rows = structure.get("hidden_rows", set())
        hidden_cols = structure.get("hidden_cols", set())
        if set(range(c0, c1 + 1)) <= hidden_cols or set(range(r0, r1 + 1)) <= hidden_rows:
            _step("A3", f"residual skipped (fully hidden) at {_a1(r0, c0, r1, c1)}")
            return None
        kind = "isolated_cell" if n_cells == 1 else "non_tabular_block"
        _step("A3", f"residual ({kind}) at {_a1(r0, c0, r1, c1)} -> summary")
        return Residual(sheet_name, bbox, kind, content=block.tolist())

    def a4_split_stacked_tables(self, candidates: list[CandidateBlock]):
        """A4 — re-split each candidate block into stacked sub-tables, top to bottom.

        Several tables can be stacked vertically, separated not by an empty row
        (otherwise A3's connected components would already have split them) but by
        a FULL-WIDTH MERGED ROW: its text is a title (first merged row of a group)
        followed by context (subsequent merged rows). The non-merged rows below
        form the body (distinct columns), up to the next full-width merged row ->
        a new table.

        Returns (tables, residuals). Ids are assigned here, contiguous per sheet.
        """
        tables, residuals = [], []
        t_idx = 0
        for cb in candidates:
            sub_tables, sub_residuals = self._split_block(cb.sheet, cb.row_abs, cb.col_abs, cb.c0, cb.c1, cb.grid, cb.local_merges)
            for st in sub_tables:
                t_idx += 1
                st.id = f"{cb.sheet}.t{t_idx}"
                tables.append(st)
                sr0, sc0, sr1, sc1 = st.bbox
                _step(
                    "A4",
                    f"table {st.id} at {_a1(sr0, sc0, sr1, sc1)} "
                    f"({st.grid.shape[0]}×{st.grid.shape[1]})" + (f'  title="{st.title}"' if st.title else "") + (f"  +{len(st.context)} context row(s)" if st.context else ""),
                )
            residuals.extend(sub_residuals)
        return tables, residuals

    def a5_strip_leading_label_columns(self, tables: list[DetectedTable]) -> None:
        """A5 — remove the merged label columns at the START of each table.

        Vertical mirror of A4's title row: a column is a label rather than data
        when a single vertical merge covers at least STRIP_LABEL_COLUMN_HEIGHT_RATIO
        of its useful height (occupied rows, ignoring trailing empties that belong
        to taller neighbours). Scanning left to right, each label column's value
        moves to the context and the column is removed; the scan stops at the first
        data column. Modifies in place.
        """
        for t in tables:
            h, w = t.grid.shape
            if h == 0 or w == 0:
                _step("A5", f"{t.id} : empty table, skipped")
                continue
            ratio = self.STRIP_LABEL_COLUMN_HEIGHT_RATIO
            dc0_sheet = t.data_bbox[1] if t.data_bbox is not None else 0  # sheet col. of grid[:,0]

            # Strip a left-to-right run of label columns; stop at the first column
            # with no qualifying vertical merge (the data). Useful height = rows
            # covered by a merge or non-empty (trailing empties ignored).
            n_strip = 0
            for j in range(w):
                col_letter = get_column_letter(dc0_sheet + j + 1)
                # Gate: vertical merges (>=2 rows) covering the column. Horizontal
                # header merges do not count.
                vmerges = [(a, b, c, d) for (a, b, c, d) in t.local_merges if b <= j <= d and (c - a + 1) >= 2]
                if not vmerges:
                    _step("A5", f"[ANALYSIS] {t.id} column {col_letter} : no vertical merge -> data column (stop)")
                    break
                # useful height = occupied rows (covered by a merge OR non-empty);
                # tallest merge = the SINGLE candidate merge.
                merged_rows: set[int] = set()
                for a, b, c, d in vmerges:
                    merged_rows.update(range(a, c + 1))
                used = sum(1 for r in range(h) if r in merged_rows or _nonempty(t.grid[r, j]))
                biggest = max(c - a + 1 for (a, b, c, d) in vmerges)
                is_label = biggest >= ratio * used
                _step(
                    "A5",
                    f"[ANALYSIS] {t.id} column {col_letter} : "
                    f"max merge {biggest}/{used} useful row(s) = "
                    f"{biggest / used:.0%}" + (f" >={ratio:.0%} -> label" if is_label else f" <{ratio:.0%} -> data column (stop)"),
                )
                if not is_label:
                    break  # first data column -> stop
                n_strip += 1
            if n_strip == 0:
                _step("A5", f"{t.id} : no leading label column (single merge >={ratio:.0%} of the useful height) -> unchanged")
                continue

            _step("A5", f"{t.id} : {n_strip} leading label column(s) detected (single merge >={ratio:.0%} of the useful height) -> removal + context")

            # value(s) of each removed column -> context (nothing lost: all
            # non-empty values, top to bottom, joined with " | ").
            for col in range(n_strip):
                parts = [str(v).strip() for v in t.grid[:, col] if _nonempty(v) and str(v).strip()]
                col_letter = get_column_letter(dc0_sheet + col + 1)
                if parts:
                    val = " | ".join(parts)
                    t.context.append(val)
                    _step("A5", f'[LABEL] {t.id} column {col_letter} -> context "{val}"')
                else:
                    _step("A5", f"[LABEL] {t.id} column {col_letter} empty -> removed without value")

            # physically remove the n_strip left columns + remap the merges.
            n_merges_before = len(t.local_merges)
            t.grid = t.grid[:, n_strip:]
            t.local_merges = [(a, max(b - n_strip, 0), c, d - n_strip) for (a, b, c, d) in t.local_merges if d >= n_strip]  # ignore fully removed merges
            if t.data_bbox is not None:
                dr0, dc0, dr1, dc1 = t.data_bbox
                t.data_bbox = (dr0, dc0 + n_strip, dr1, dc1)  # left data edge shifted

            _step("A5", f"{t.id} : {n_strip} column(s) removed, {n_merges_before - len(t.local_merges)} merge(s) deleted -> body {t.grid.shape[0]}×{t.grid.shape[1]}, data {t.plage_donnees}")

    def _split_block(self, sheet_name: str, row_abs: list[int], col_abs: list[int], c0: int, c1: int, block: np.ndarray, local_merges: list):
        """Splitting algorithm for ONE block (see a4_split_stacked_tables).

        `row_abs[i]`/`col_abs[j]` map cell (i, j) of the block to its absolute
        sheet row/column (the bboxes stay correct even if rows/columns were
        removed). `c0`/`c1` are the ORIGINAL columns of the block (full footprint);
        `col_abs` gives the columns actually KEPT.
        Returns (sub_tables, residuals). The sub-tables have a provisional `id`,
        renumbered by A4.
        """
        h, w = block.shape

        # 1) sep[r]: row r is a separator if covered by a merge at least
        #    SPLIT_SEPARATOR_WIDTH_RATIO of the block width wide.
        #    anchor_val[r]: the title/context value if r is a merge anchor.
        sep = [False] * h
        anchor_val: dict[int, Any] = {}
        min_sep_cols = self.SPLIT_SEPARATOR_WIDTH_RATIO * w
        for mr0, mc0, mr1, mc1 in local_merges:
            if (mc1 - mc0 + 1) >= min_sep_cols:  # wide enough to separate
                for r in range(max(mr0, 0), min(mr1 + 1, h)):
                    sep[r] = True

        # 1b) value of each separator row: join ALL its non-empty values left to
        #     right with " | " so nothing is lost. Merged cells are empty except
        #     their anchor, so each merge is read once.
        for r in range(h):
            if not sep[r]:
                continue
            parts = [str(v).strip() for v in block[r, :] if _nonempty(v) and str(v).strip()]
            if parts:
                anchor_val[r] = " | ".join(parts)

        # 2) segmentation into consecutive runs of the same type (sep / body)
        segments = []  # (type, start, end) inclusive
        i = 0
        while i < h:
            j = i
            while j + 1 < h and sep[j + 1] == sep[i]:
                j += 1
            segments.append(("sep" if sep[i] else "body", i, j))
            i = j + 1

        # 3) reassembly: [sep]? + body = one table
        tables, residuals = [], []
        pending_title: Optional[str] = None
        pending_context: list[str] = []
        pending_start: Optional[int] = None  # first row (block) of the title segment

        def emit(body_s: int, body_e: int, title, context, seg_start):
            top = seg_start if seg_start is not None else body_s
            r0_abs, r1_abs = row_abs[top], row_abs[body_e]
            n_cells = _count_nonempty(block[body_s : body_e + 1, :])
            bh, bw = body_e - body_s + 1, w
            if not _is_table_area(bh, bw, n_cells):  # body too thin
                if self.keep_split_residuals:
                    residuals.append(Residual(sheet_name, (r0_abs, c0, r1_abs, c1), "non_tabular_block", content=block[body_s : body_e + 1, :].tolist()))
                dest = "-> residual" if self.keep_split_residuals else "(dropped, not kept)"
                _step("A4", f"[SPLIT] body too thin at {_a1(r0_abs, c0, r1_abs, c1)} {dest}")
                return
            body_grid = block[body_s : body_e + 1, :]
            body_merges = [(a - body_s, b, c - body_s, d) for (a, b, c, d) in local_merges if a >= body_s and c <= body_e]
            # KEPT range: body only (excluding title/context) over the columns actually kept
            data_bbox = (row_abs[body_s], col_abs[0], row_abs[body_e], col_abs[-1])
            tables.append(DetectedTable(f"{sheet_name}.tmp", sheet_name, (r0_abs, c0, r1_abs, c1), body_grid.copy(), body_merges, title=title, context=list(context), data_bbox=data_bbox))

        for typ, s, e in segments:
            if typ == "sep":
                # title = first anchor value of the group, context = the rest
                vals = [anchor_val[r] for r in range(s, e + 1) if r in anchor_val]
                if pending_title is None and pending_start is None:
                    pending_start = s
                if vals:
                    if pending_title is None:
                        pending_title = vals[0]
                        pending_context = vals[1:]
                    else:
                        pending_context += vals
            else:  # body
                emit(s, e, pending_title, pending_context, pending_start)
                pending_title, pending_context, pending_start = None, [], None

        # trailing title segment with no body -> free title/note
        if pending_title is not None or pending_start is not None:
            r_abs = row_abs[pending_start] if pending_start is not None else row_abs[-1]
            if self.keep_split_residuals:
                residuals.append(Residual(sheet_name, (r_abs, c0, r_abs, c1), "free_title", content=[pending_title] + pending_context if pending_title else None))
            dest = "-> residual" if self.keep_split_residuals else "(dropped, not kept)"
            _step("A4", f"[SPLIT] orphan title at {_a1(r_abs, c0, r_abs, c1)} {dest}")

        return tables, residuals

    @staticmethod
    def _connected_components(mask: np.ndarray):
        """4-connectivity; an empty row/column breaks connectivity -> distinct blocks."""
        visited = np.zeros_like(mask, dtype=bool)
        rows, cols = mask.shape
        comps = []
        for r in range(rows):
            for c in range(cols):
                if mask[r, c] and not visited[r, c]:
                    stack, cells = [(r, c)], []
                    visited[r, c] = True
                    while stack:
                        y, x = stack.pop()
                        cells.append((y, x))
                        for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                            ny, nx = y + dy, x + dx
                            if 0 <= ny < rows and 0 <= nx < cols and mask[ny, nx] and not visited[ny, nx]:
                                visited[ny, nx] = True
                                stack.append((ny, nx))
                    ys = [p[0] for p in cells]
                    xs = [p[1] for p in cells]
                    comps.append((min(ys), min(xs), max(ys), max(xs)))
        # NB: a table cut by a fully empty internal row would be split in two here
        #     (known limitation, see the spec). Accepted to stay generic.
        return comps

    @staticmethod
    def _coverage(grid, tables, residuals):
        total = _count_nonempty(grid)
        if total == 0:
            return 1.0
        covered = 0
        for t in tables:
            r0, c0, r1, c1 = t.bbox
            covered += _count_nonempty(grid[r0 : r1 + 1, c0 : c1 + 1])
        return round(covered / total, 2)

    # ================================================================== #
    # PHASE B — run per table
    # ================================================================== #
    def b1_detect_orientation(self, table: DetectedTable) -> None:
        """B1 — numeric headers + textual first column => transposed or cross-tab."""
        g = table.grid
        if g.shape[0] < 2 or g.shape[1] < 2:
            return
        head = [v for v in g[0, 1:] if _nonempty(v)]
        col0 = [v for v in g[1:, 0] if _nonempty(v)]
        head_num = sum(_is_data_like(v) for v in head) / max(len(head), 1)
        col0_txt = sum(not _is_data_like(v) for v in col0) / max(len(col0), 1)
        if head_num > self.ORIENT_HEADER_NUMERIC_RATIO and col0_txt > self.ORIENT_FIRST_COL_TEXT_RATIO:
            inner = g[1:, 1:]
            inner_num = np.vectorize(lambda v: _is_data_like(v) or not _nonempty(v))(inner).mean()
            if inner_num > self.ORIENT_INNER_NUMERIC_RATIO and g.shape[0] <= g.shape[1] + 1:
                table.orientation = "transposed"
                table.grid = g.T.copy()
                table.local_merges = [(b, a, d, c) for (a, b, c, d) in table.local_merges]
                _step("B1", f"{table.id} — transposed orientation -> straightened (df.T)")
                return
            table.orientation = "cross-tab"
            _step("B1", f"{table.id} — cross-tab table")
            return
        _step("B1", f"{table.id} — normal orientation")

    def b2_to_dataframe(self, table: DetectedTable) -> None:
        """B2 — auto-fill the merges then convert to a pandas DataFrame.

        1. AUTO-FILL: each merged range receives its anchor value across its
           whole extent.
        2. CONVERSION: the first row becomes the header, the rest the data;
           empty/duplicate column names are cleaned and made unique.
        """
        g = table.grid
        if g.shape[0] < 2:
            raise ValueError("table without a data row")

        # 1) auto-fill the merged cells from their anchor value
        n = 0
        for r0, c0, r1, c1 in table.local_merges:
            if not (0 <= r0 < g.shape[0] and 0 <= c0 < g.shape[1]):
                continue
            anchor = g[r0, c0]
            for rr in range(r0, min(r1 + 1, g.shape[0])):
                for cc in range(c0, min(c1 + 1, g.shape[1])):
                    if not _nonempty(g[rr, cc]):
                        g[rr, cc] = anchor
                        n += 1

        # 2) pandas conversion: first row = header, unique names
        header: list[str] = []
        seen: dict[str, int] = {}
        any_named = False
        unnamed: list[str] = []  # header names of the title-less columns
        for i, v in enumerate(g[0]):
            if _nonempty(v):
                # Single-line header: flatten any line break / multiple spaces into
                # a single space, for all tables.
                name = " ".join(str(v).split())
                any_named = True
            else:
                name = f"col_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}.{seen[name]}"
            else:
                seen[name] = 0
            header.append(name)
            if not _nonempty(v):
                unnamed.append(name)
        # No cell of the first row supplied a label -> auto-generated header.
        table.has_column_names = any_named

        df = pd.DataFrame(g[1:], columns=pd.Index(header))
        # B2 (option): drop the columns whose first row supplied no title.
        if not self.keep_headerless_columns and unnamed:
            df = df.drop(columns=unnamed)
            _step("B2", f"{table.id} — {len(unnamed)} title-less column(s) dropped: {unnamed}")
        table.df = df
        _step("B2", f"{table.id} — {n} cell(s) auto-filled, DataFrame {df.shape}")
        _step("B2", f"{table.id} — columns={list(df.columns)}")

    def b3_check_empty(self, table: DetectedTable) -> None:
        """B3 — drop empty rows / columns + state ("empty" / "non-empty").

        Depending on the constructor options:
          `drop_empty_rows` -> drop any row where ALL cells are empty.
          `drop_empty_cols` -> drop any column where ALL cells are empty.
        The possible dropping of a table reduced to a single column
        (`keep_single_column_tables`) is handled by the orchestrator, after this step.
        """
        df = table.df
        n_rows_dropped = n_cols_dropped = 0
        if df is not None and not df.empty:
            if self.drop_empty_rows:
                keep = df.apply(lambda row: any(_nonempty(v) for v in row), axis=1)
                n_rows_dropped = int((~keep).sum())
                df = cast(pd.DataFrame, df[keep]).reset_index(drop=True)
            if self.drop_empty_cols:
                keep_cols = [c for c in df.columns if any(_nonempty(v) for v in df[c])]
                n_cols_dropped = df.shape[1] - len(keep_cols)
                df = cast(pd.DataFrame, df[keep_cols])
            table.df = df
        has_data = df is not None and any(_nonempty(v) for v in df.to_numpy().ravel())
        table.etat = "non-empty" if has_data else "empty"
        _step("B3", f"{table.id} — {n_rows_dropped} empty row(s) and {n_cols_dropped} empty column(s) dropped, state: {table.etat}")

    def b4_clean_and_coerce(self, table: DetectedTable) -> None:
        """B4 — column recognition: dates, numeric, text.

        A numeric column is RECOGNISED but its values are NOT converted: only the
        whitespace grouping thousands is stripped (currency, %, decimal comma are
        kept, the value stays a string).

        Can be disabled with `coerce_types=False`: the columns then stay as raw
        values (no typing, no conversion), exactly as read from the sheet."""
        df = table.df
        if df is None:
            raise ValueError("table without a DataFrame (B2 not run)")
        if not self.coerce_types:
            _step("B4", f"{table.id} — typing disabled (coerce_types=False), raw data kept")
            return
        report = []
        for col in df.columns:
            s = df[col]
            name = str(col)

            # 1) Already dates (openpyxl returns datetimes for date-formatted cells)
            if any(isinstance(v, (dt.date, dt.datetime)) for v in s if _nonempty(v)):
                parsed_dt = pd.to_datetime(s, errors="coerce")
                df[col] = pd.Series([v if pd.notna(v) else None for v in parsed_dt], index=s.index, dtype=object)
                report.append(f"{name}=date")
                continue

            # 2) Numeric: column RECOGNISED but values NOT converted — only the
            #    thousands-grouping whitespace is stripped (currency, %, decimal
            #    comma kept). Non-numeric values in the column are left as-is.
            non_null = sum(1 for v in s if _nonempty(v))
            numeric = sum(1 for v in s if _nonempty(v) and (_is_number(v) or _looks_numeric(v)))
            if non_null and numeric / non_null >= 0.7:
                df[col] = s.map(lambda v: (v if _is_number(v) else _despace_thousands(v)) if _nonempty(v) else None)
                report.append(f"{name}=num")
                continue

            # 3) Dates stored as text. format="mixed" parses each value on its own
            #    (formats are heterogeneous) without the "Could not infer format"
            #    warning.
            dt_parsed = pd.to_datetime(s, errors="coerce", dayfirst=True, format="mixed")
            if non_null and dt_parsed.notna().sum() / non_null >= 0.7:
                df[col] = pd.Series([v if pd.notna(v) else None for v in dt_parsed], index=s.index, dtype=object)
                report.append(f"{name}=date(txt)")
                continue

            # 4) Otherwise: cleaned text (raw data kept, including 0)
            df[col] = s.map(lambda v: str(v).strip() if _nonempty(v) else None)
            report.append(f"{name}=str")

        table.df = df
        _step("B4", f"{table.id} — types : {', '.join(report)}")

    def b5_validate_and_tag(self, table: DetectedTable) -> None:
        """B5 — provenance attached."""
        df = table.df
        if df is None:
            raise ValueError("table without a DataFrame (B2 not run)")
        table.provenance = {
            "file": self.path,
            "sheet": table.sheet,
            "range": table.plage,
            "data_range_used": table.plage_donnees,
            "title": table.title,
            "context": table.context,
            "orientation": table.orientation,
            "state": table.etat,
        }
        df.attrs["provenance"] = table.provenance
        _step("B5", f"{table.id} — provenance attached, {df.shape}, status={table.status}")
